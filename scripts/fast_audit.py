import sys
import openpyxl

def fast_audit():
    f_oscar = 'oscar_satellite_data_full_perfection.xlsx'
    f_ceos = 'satellite_data_full.xlsx'
    f_comb = 'combined_satellite_data_strict.xlsx'

    print("Loading file row counts...")
    wb_o = openpyxl.load_workbook(f_oscar, read_only=True)
    ws_o = wb_o.active
    print(f"OSCAR rows: {ws_o.max_row}")
    # Get column names
    o_cols = [c.value for c in next(ws_o.iter_rows(min_row=1, max_row=1))]

    wb_c = openpyxl.load_workbook(f_ceos, read_only=True)
    ws_c = wb_c.active
    print(f"CEOS rows: {ws_c.max_row}")
    
    wb_comb = openpyxl.load_workbook(f_comb, read_only=True)
    ws_comb = wb_comb.active
    print(f"Combined rows: {ws_comb.max_row}")
    
    comb_cols = [c.value for c in next(ws_comb.iter_rows(min_row=1, max_row=1))]
    print(f"Combined columns: {len(comb_cols)}")
    
    dup_cols = set([x for x in comb_cols if comb_cols.count(x) > 1 and x is not None])
    print(f"Duplicate column names: {list(dup_cols)}")
    
    bad_cols = [str(c) for c in comb_cols if str(c).endswith('_ceos') or str(c).endswith('_oscar') or str(c).endswith('_x') or str(c).endswith('_y')]
    print(f"Columns with internal merge suffixes left: {bad_cols}")
    
    # Try reading the combined rows completely using read_only to see if it's very fast
    print("Reading combined rows...")
    hashes = set()
    dup_row_count = 0
    row_idx = 0
    c_names = []
    o_names = []
    
    # Find indices for Sat_Full_Name and Sat_Acronym
    idx_full = comb_cols.index('Sat_Full_Name') if 'Sat_Full_Name' in comb_cols else -1
    idx_acronym = comb_cols.index('Sat_Acronym') if 'Sat_Acronym' in comb_cols else -1

    matched = 0

    iteration = ws_comb.iter_rows(min_row=2, values_only=True)
    for row in iteration:
        row_idx += 1
        if len(row) == 0 or all(cell is None for cell in row):
            continue
        
        # Check duplicate row
        r_tuple = tuple(str(x) for x in row)
        if r_tuple in hashes:
            dup_row_count += 1
        else:
            hashes.add(r_tuple)
            
        full_name = row[idx_full] if idx_full != -1 else None
        acronym = row[idx_acronym] if idx_acronym != -1 else None
        
        if full_name and acronym:
            matched += 1
            
        if full_name:
            c_names.append(full_name)
        if acronym:
            o_names.append(acronym)

    print(f"Duplicate exact rows in combined: {dup_row_count}")
    print(f"Rows with both full name (CEOS) and acronym (OSCAR) populated: {matched}")
    
if __name__ == '__main__':
    fast_audit()
